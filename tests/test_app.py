"""Pruebas de arranque y navegación de la TUI."""

from __future__ import annotations

import os

os.environ["GPROJECT_DB"] = "memory"

import pytest

from textual.widgets import Button, Input, Select

from gproject.app import GProjectApp
from gproject.modals.filter_bar import FilterBar
from gproject.modals.task_form import TaskForm
from gproject.screens.main import HelpModal, MainScreen
from gproject.widgets.gantt_chart import GanttView
from gproject.widgets.kanban_board import KanbanView


async def test_arranque_con_seed():
    app = GProjectApp(seed=True)
    async with app.run_test(size=(120, 30)) as pilot:
        await pilot.pause()
        assert isinstance(app.screen, MainScreen)
        gantt = app.screen.query_one(GanttView)
        assert len(gantt.filas) == 11
        assert app.proyecto_activo_id is not None


async def test_navegacion_y_zoom():
    app = GProjectApp(seed=True)
    async with app.run_test(size=(120, 30)) as pilot:
        await pilot.pause()
        gantt = app.screen.query_one(GanttView)
        await pilot.press("down", "down")
        await pilot.pause()
        assert gantt.cursor == 2
        nivel_inicial = gantt.nivel
        await pilot.press("minus")
        await pilot.pause()
        assert gantt.nivel == nivel_inicial + 1


async def test_ciclar_zoom_con_z():
    app = GProjectApp(seed=True)
    async with app.run_test(size=(120, 30)) as pilot:
        await pilot.pause()
        gantt = app.screen.query_one(GanttView)
        assert gantt.nivel == 0 and gantt.nivel_zoom_label == "Día"
        await pilot.press("z")
        await pilot.pause()
        assert gantt.nivel == 1 and gantt.nivel_zoom_label == "Semana"
        await pilot.press("z", "z")  # Mes → vuelve a Día
        await pilot.pause()
        assert gantt.nivel == 0


async def test_abrir_y_cerrar_modales():
    app = GProjectApp(seed=True)
    async with app.run_test(size=(120, 30)) as pilot:
        await pilot.pause()
        await pilot.press("n")
        await pilot.pause()
        assert isinstance(app.screen, TaskForm)
        await pilot.press("escape")
        await pilot.pause()
        assert isinstance(app.screen, MainScreen)

        await pilot.press("question_mark")
        await pilot.pause()
        assert isinstance(app.screen, HelpModal)
        await pilot.press("escape")
        await pilot.pause()
        assert isinstance(app.screen, MainScreen)


async def test_crear_proyecto_y_tarea_sin_seed():
    """Regresión: el formulario de tarea no debe cerrar la app en un proyecto vacío."""
    app = GProjectApp(seed=False)
    async with app.run_test(size=(120, 30)) as pilot:
        await pilot.pause()
        # Crear proyecto
        await pilot.press("N")
        await pilot.pause()
        app.screen.query_one("#nombre", Input).value = "Proyecto X"
        app.screen.query_one("#guardar", Button).press()
        await pilot.pause()
        assert app.proyecto_activo_id is not None

        # Crear tarea en proyecto vacío (campo «padre» sin opciones)
        await pilot.press("n")
        await pilot.pause()
        assert isinstance(app.screen, TaskForm)
        app.screen.query_one("#titulo", Input).value = "Tarea 1"
        app.screen.query_one("#guardar", Button).press()
        await pilot.pause()
        assert isinstance(app.screen, MainScreen)
        gantt = app.screen.query_one(GanttView)
        assert len(gantt.filas) == 1
        assert gantt.filas[0].tarea.titulo == "Tarea 1"


async def test_abrir_filtro():
    app = GProjectApp(seed=True)
    async with app.run_test(size=(120, 30)) as pilot:
        await pilot.pause()
        await pilot.press("slash")
        await pilot.pause()
        assert isinstance(app.screen, FilterBar)
        await pilot.press("escape")
        await pilot.pause()
        assert isinstance(app.screen, MainScreen)


async def test_cambiar_estado_persiste():
    app = GProjectApp(seed=True)
    async with app.run_test(size=(120, 30)) as pilot:
        await pilot.pause()
        gantt = app.screen.query_one(GanttView)
        # Ir a una hoja conocida y cambiar su estado con espacio
        gantt.seleccionar_tarea(
            next(f.tarea.id for f in gantt.filas if f.tarea.titulo == "Frontend")
        )
        await pilot.pause()
        tarea = gantt.tarea_actual
        estado_previo = tarea.estado
        await pilot.press("space")
        await pilot.pause()
        actualizada = app.db.get_tarea(tarea.id)
        assert actualizada.estado != estado_previo


async def test_alternar_a_vista_kanban():
    app = GProjectApp(seed=True)
    async with app.run_test(size=(120, 30)) as pilot:
        await pilot.pause()
        assert app.screen.vista == "gantt"
        assert app.screen.query_one(GanttView).display

        await pilot.press("v")
        await pilot.pause()

        assert app.screen.vista == "kanban"
        assert app.screen.query_one(KanbanView).display
        assert not app.screen.query_one(GanttView).display


async def test_exportar_csv_crea_archivo(tmp_path, monkeypatch):
    """Pulsar 'x' abre el modal; al exportar en CSV crea el archivo con encabezados."""
    monkeypatch.chdir(tmp_path)
    app = GProjectApp(seed=True)
    async with app.run_test(size=(120, 30)) as pilot:
        await pilot.pause()
        await pilot.press("x")
        await pilot.pause()
        from gproject.modals.export_form import ExportModal

        assert isinstance(app.screen, ExportModal)
        app.screen.query_one("#exportar", Button).press()
        await pilot.pause()

    csvs = list(tmp_path.glob("*_tareas_*.csv"))
    assert len(csvs) == 1
    contenido = csvs[0].read_text(encoding="utf-8-sig")
    lineas = contenido.splitlines()
    assert lineas[0].startswith("ID,Tarea,Nivel")
    assert len(lineas) == 12  # encabezado + 11 tareas del seed


async def test_exportar_xlsx_crea_archivo(tmp_path, monkeypatch):
    """Elegir Excel en el modal genera un .xlsx legible con una fila por tarea."""
    monkeypatch.chdir(tmp_path)
    app = GProjectApp(seed=True)
    async with app.run_test(size=(120, 30)) as pilot:
        await pilot.pause()
        await pilot.press("x")
        await pilot.pause()
        app.screen.query_one("#formato", Select).value = "xlsx"
        await pilot.pause()
        app.screen.query_one("#exportar", Button).press()
        await pilot.pause()

    xlsxs = list(tmp_path.glob("*_tareas_*.xlsx"))
    assert len(xlsxs) == 1
    from openpyxl import load_workbook

    ws = load_workbook(xlsxs[0]).active
    assert ws["A1"].value == "ID" and ws["B1"].value == "Tarea"
    assert ws.max_row == 12  # encabezado + 11 tareas


def test_exportar_csv_orden_jerarquico(tmp_path):
    """exportar_csv ordena padres antes que hijos y marca el nivel."""
    import csv as _csv

    from gproject.db import Database
    from gproject import export

    db = Database(":memory:")
    pid = db.create_proyecto("P")
    a = db.create_tarea(pid, "A")
    db.create_tarea(pid, "A.1", parent_id=a)
    db.create_tarea(pid, "B")
    ruta = export.exportar_csv(db.list_tareas(pid), db.list_dependencias(pid),
                               tmp_path / "t.csv")
    db.close()
    filas = list(_csv.reader(ruta.open(encoding="utf-8-sig")))
    titulos = [(f[1], f[2]) for f in filas[1:]]  # (Tarea, Nivel)
    assert titulos == [("A", "0"), ("A.1", "1"), ("B", "0")]


async def test_kanban_cambia_estado_con_espacio():
    app = GProjectApp(seed=True)
    async with app.run_test(size=(120, 30)) as pilot:
        await pilot.pause()
        await pilot.press("v")
        await pilot.pause()

        kanban = app.screen.query_one(KanbanView)
        frontend_id = next(t.id for t in kanban.por_estado["todo"] if t.titulo == "Frontend")
        kanban.seleccionar_tarea(frontend_id)
        await pilot.pause()

        assert kanban.tarea_actual.estado == "todo"
        await pilot.press("space")
        await pilot.pause()

        actualizada = app.db.get_tarea(frontend_id)
        assert actualizada.estado == "doing"
        assert app.screen.query_one(KanbanView).tarea_actual.id == frontend_id
