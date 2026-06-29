"""Exportación de las tareas de un proyecto a CSV o Excel (.xlsx).

La tabla respeta el orden jerárquico del árbol (padres antes que hijos) e incluye
una columna ``Nivel`` con la profundidad para reconstruir la jerarquía.
"""

from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path

from gproject.models import Dependencia, Tarea

ESTADO_LABEL = {"todo": "Pendiente", "doing": "En progreso", "done": "Hecho"}

ENCABEZADOS = [
    "ID", "Tarea", "Nivel", "Estado", "Prioridad", "Inicio", "Fin",
    "Duración (días)", "Progreso (%)", "Hito", "Depende de", "Etiquetas",
]


def nombre_archivo(nombre_proyecto: str, ext: str) -> str:
    """Devuelve ``<slug>_tareas_<YYYYMMDD-HHMMSS>.<ext>``."""
    slug = re.sub(r"[^\w-]+", "_", nombre_proyecto.strip().lower()).strip("_") or "proyecto"
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"{slug}_tareas_{ts}.{ext}"


def _ordenar_jerarquico(tareas: list[Tarea]) -> list[tuple[Tarea, int]]:
    """Devuelve ``[(tarea, nivel)]`` en orden de árbol (DFS por orden, id)."""
    hijos: dict[int | None, list[Tarea]] = {}
    for t in tareas:
        hijos.setdefault(t.parent_id, []).append(t)
    for lst in hijos.values():
        lst.sort(key=lambda t: (t.orden, t.id))

    orden: list[tuple[Tarea, int]] = []

    def recorrer(parent_id: int | None, nivel: int) -> None:
        for t in hijos.get(parent_id, []):
            orden.append((t, nivel))
            recorrer(t.id, nivel + 1)

    recorrer(None, 0)
    return orden


def _filas(tareas: list[Tarea], dependencias: list[Dependencia]) -> list[list]:
    deps_por_tarea: dict[int, list[int]] = {}
    for d in dependencias:
        deps_por_tarea.setdefault(d.tarea_id, []).append(d.depende_de_id)

    filas: list[list] = []
    for t, nivel in _ordenar_jerarquico(tareas):
        filas.append([
            t.id,
            t.titulo,
            nivel,
            ESTADO_LABEL.get(t.estado, t.estado),
            t.prioridad,
            t.fecha_inicio.isoformat() if t.fecha_inicio else "",
            t.fecha_fin.isoformat() if t.fecha_fin else "",
            "" if t.es_hito else t.duracion_dias,
            "" if t.es_hito else t.progreso,
            "Sí" if t.es_hito else "No",
            " ".join(str(x) for x in deps_por_tarea.get(t.id, [])),
            ", ".join(e.nombre for e in t.etiquetas),
        ])
    return filas


def exportar_csv(
    tareas: list[Tarea], dependencias: list[Dependencia], ruta: Path | str
) -> Path:
    """Escribe las tareas a un CSV (UTF-8 con BOM para que Excel lea los acentos)."""
    ruta = Path(ruta)
    with ruta.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(ENCABEZADOS)
        writer.writerows(_filas(tareas, dependencias))
    return ruta


def exportar_xlsx(
    tareas: list[Tarea], dependencias: list[Dependencia], ruta: Path | str
) -> Path:
    """Escribe las tareas a un .xlsx con encabezados en negrita y filtros."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    ruta = Path(ruta)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tareas"

    ws.append(ENCABEZADOS)
    for celda in ws[1]:
        celda.font = Font(bold=True)

    filas = _filas(tareas, dependencias)
    for fila in filas:
        ws.append(fila)

    # Anchos de columna aproximados al contenido + autofiltro y panel fijo.
    for col, enc in enumerate(ENCABEZADOS, start=1):
        ancho = max([len(enc)] + [len(str(f[col - 1])) for f in filas] + [4])
        ws.column_dimensions[get_column_letter(col)].width = min(ancho + 2, 50)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ENCABEZADOS))}{len(filas) + 1}"
    ws.freeze_panes = "A2"

    wb.save(ruta)
    return ruta
